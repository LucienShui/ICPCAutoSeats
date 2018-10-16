import openpyxl
import random
width = 24  # 场地的宽度(单位:队)
length = 16  # 场地的长度(单位:队)
tot = 370  # 总共有多少队
minimal_accepted_distance = 100  # 同一学校的任意两个队伍之间可接受的最小曼哈顿距离
path = "xlsx.xlsx"  # excel表格的路径, 第一行不能包含队伍(因为程序会跳过第一行), 表格中第一列留空，第二列为学校名


def get_coordinate(index):  # 根据index返回高斯坐标
    index = index - 1
    return index // width + 1, index % width + 1


def get_alpha(index):  # 根据index返回实际的座位号
    index = index - 1
    block_size = (width // 2) * (length // 2)
    belong = index // block_size
    return "%c%02d" % ("ABCD"[belong], index % block_size + 1)


def process():  # 为每个学校随机分配一个index, 范围:[1, tot], 保证唯一
    vis = []
    flag = True
    workbook = openpyxl.load_workbook(path)
    sheet = workbook.active
    for row in sheet.rows:
        if flag:
            flag = False
            continue
        tmp = random.randint(1, tot)
        while tmp in vis:
            tmp = random.randint(1, tot)
        vis.append(tmp)
        row[0].value = tmp
    workbook.save("new.xlsx")


def min_dist(array):  # 返回array中任意两点间的最小距离
    array_len = array.__len__()
    minimal = tot << 2
    if array_len == 1:
        return minimal
    for i in range(0, array_len):
        for j in range(i + 1, array_len):
            ux, uy = get_coordinate(array[i])
            vx, vy = get_coordinate(array[j])
            minimal = min(abs(ux - vx) + abs(uy - vy))
    return minimal


def check():  # 返回当前的一个生成是否合法
    workbook = openpyxl.load_workbook("new.xlsx")
    sheet = workbook.active
    sets = []
    cur = 0
    flag = False
    for row in sheet.rows:
        if flag:
            flag = False
            continue
        index = row[0].value
        school_name = row[1].value
        if school_name != cur:
            if min_dist(sets) <= 3:
                return False
            sets = []
        sets.append(index)
    if min_dist(sets) <= minimal_accepted_distance:
        return False
    return True


def show(filename):  # 展示表中的内容
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active
    for row in sheet.rows:
        for cell in row:
            print(cell.value, end=" | ")
        print()


def trans():  # 将合法表中的index转换为座位号
    workbook = openpyxl.load_workbook("new.xlsx")
    sheet = workbook.active
    flag = True
    for row in sheet.rows:
        if flag:
            flag = False
            continue
        row[0].value = get_alpha(row[0].value)
    workbook.save("new.xlsx")


if __name__ == '__main__':
    process()
    while not check():
        process()
    trans()
